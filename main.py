import cv2
import numpy as np
import face_recognition
import os
import xlrd, xlwt
from xlutils.copy import copy as xl_copy
from datetime import date, datetime
from sklearn.metrics import accuracy_score
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# from PIL import ImageGrab

path = 'Training_images'
images = []
classNames = []
myList = os.listdir(path)
# print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
# print(classNames)


def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList
encodeListKnown = findEncodings(images)
print('Face Encoding Complete')
cap = cv2.VideoCapture(0)
wb = load_workbook('attendence_excel1.xlsx')
op='Y'
already_attendence_taken = ""
sheet = input('Please give current subject lecture name: ')
print(sheet)
sheetlist = wb.sheetnames
if sheet in sheetlist:
    ws=wb[sheet]
    list = []
    rows=ws.max_row
    cols=ws.max_column
    newcol=cols+1
    ws.insert_cols(newcol)
    for i in range(2,rows+1):
        list.append(ws['A' + str(i)].value)
    for i in range(2,rows+1):
        ws.cell(row=i,column=newcol).value = 'Absent'
    while True:
        success, img = cap.read()
        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                while(op=='Y' or op == 'y'):
                    ws.cell(row=1, column=newcol).value = date.today()
                    if((already_attendence_taken != name) and (name != "Unknown")):
                        if name in list:   
                            name_row = list.index(name) +2
                            ws.cell(row=name_row, column=newcol).value = 'Present'
                            print("attendence taken")
                            wb.save('attendence_excel1.xlsx')
                            already_attendence_taken = name
                            op = input('Do you want to continue: (Y/N) ')
                    else:
                        print("next student")
                        op = input('Do you want to continue: (Y/N) ')
                    break
        cv2.imshow('Webcam', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
else:
    print("Lecture Not available.")
    add= (input('Do you want to this to add thi course:(Y/N)'))
    if(add == 'Y' or add == 'y'):
        ws=wb.create_sheet(title=sheet)
        list = []
        ws.title = "Attendence"
        ws['A1'] = "Name\Date"
        print()
        student= int(input('Please Enter the number of the Student who have enrolled for this subject:'))
        print("Please Enter the names of the student")
        for i in range(1,student+1):
            sname = input()
            ws.append([sname])
        rows=ws.max_row
        cols=ws.max_column
        newcol=cols+1
        ws.insert_cols(newcol)
        for i in range(2,rows+1):
            list.append(ws['A' + str(i)].value)
        print(list)
        for i in range(2,rows+1):
            ws.cell(row=i,column=newcol).value = 'Absent'
        while True:
            success, img = cap.read()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

            facesCurFrame = face_recognition.face_locations(imgS)
            encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

            for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
                matchIndex = np.argmin(faceDis)

                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    while(op=='Y' or op == 'y'):
                        ws.cell(row=1, column=newcol).value = date.today()
                        if((already_attendence_taken != name) and (name != "Unknown")):
                            if name in list:   
                                name_row = list.index(name) +2
                                ws.cell(row=name_row, column=newcol).value = 'Present'
                                print("attendence taken")
                                wb.save('attendence_excel1.xlsx')
                                already_attendence_taken = name
                                op = input('Do you want to continue: (Y/N) ')
                        else:
                            print("next student")
                            op = input('Do you want to continue: (Y/N) ')
                        break
            cv2.imshow('Webcam', img)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
